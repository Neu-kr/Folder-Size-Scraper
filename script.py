import os
from openpyxl import Workbook, load_workbook

def get_folder_size(folder_path):
    """Recursively calculate the total size of all files in the specified folder."""
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for filename in filenames:
            file_path = os.path.join(dirpath, filename)
            if not os.path.islink(file_path):
                try:
                    total_size += os.path.getsize(file_path)
                except Exception as e:
                    print(f"Failed to get size for file {file_path}: {e}")
    return total_size

def human_readable_size(size, decimal_places=2):
    """Convert byte count to a human-readable format (e.g., KB, MB, GB)."""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size < 1024:
            return f"{size:.{decimal_places}f} {unit}"
        size /= 1024
    return f"{size:.{decimal_places}f} PB"

def scrape_folders(base_path, excel_file_path):
    subfolders = [folder for folder in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, folder))]
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Folder Name", "Capacity", "Bytes"])
    for folder in subfolders:
        folder_path = os.path.join(base_path, folder)
        size_bytes = get_folder_size(folder_path)
        size_hr = human_readable_size(size_bytes)
        ws.append([folder, size_hr, size_bytes])
        print(f"{folder}: {size_hr} ({size_bytes} bytes)")
    wb.save(excel_file_path)
    print(f"Data has been exported to {excel_file_path}.")

if __name__ == "__main__":
    base_path = input("Enter the folder path: ").strip()
    if not os.path.isdir(base_path):
        print("The entered path either does not exist or is not a directory.")
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        excel_file_path = os.path.join(script_dir, "output.xlsx")
        scrape_folders(base_path, excel_file_path)
